import openpyxl
from tkinter import *
from tkinter import filedialog as FileDialog
from io import open

ss = openpyxl.load_workbook('prueba.xlsx')
sheet = ss.get_sheet_by_name('Inventario')
# print(sheet.cell(row=1, column=5).value)
hola = 0
for i in range(1, 25):
    if sheet.cell(row = i, column = 5).value == 0:
        print(sheet.cell(row = i, column = 1).value)

root = Tk()
root.title('Moras Dulces')

Label(root, text="Inventario Moras", fg="darkblue", font=("Times New Roman",28,"bold italic")).pack()

Label(root, text="Ingredientes", fg="black", font=("Times New Roman", 20,"bold italic")).pack()
mensaje = StringVar()
mensaje.set("Bienvenida Dolores")
monitor = Label(root,textvar = mensaje, justify='left')
monitor.pack(side="left")

root.mainloop()