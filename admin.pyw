import openpyxl
from tkinter import *
from tkinter import filedialog as FileDialog
from io import open

ss = openpyxl.load_workbook('prueba.xlsx')
sheet = ss.get_sheet_by_name('Inventario')


root = Tk()
root.title('Moras Dulces')

Label(root, text="Inventario Moras", fg="darkblue", font=("Times New Roman",28,"bold italic")).pack()
for i in range(1, 25):
    if sheet.cell(row = i, column = 5).value == 0:
        Label(root, text=sheet.cell(row = i, column = 1).value, fg="black", font=("Times New Roman", 20,"bold italic")).pack()
        print(sheet.cell(row = i, column = 1).value)

mensaje = StringVar()
mensaje.set("Bienvenida Dolores")
monitor = Label(root,textvar = mensaje, justify='left')
monitor.pack(side="left")

root.mainloop()